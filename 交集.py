#!/usr/bin/python
# -*- coding: UTF-8 -*-
# 匹配excel表格里两列不同的数据
#导入xlrd模块
import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
 #读取excel文件
#括号中的字符串为你要比较的两个excel的路径，注意用“/”
wb_a = openpyxl.load_workbook('C:/Users/Administrator/Desktop/test1.xlsx')
wb_b = openpyxl.load_workbook('C:/Users/Administrator/Desktop/test2.xlsx')
#定义一个方法来获取表格中某一列的内容，返回一个列表
#在这里，我的表格中：IP是具有唯一性的，所以我用它来区分数据的差异，而IP这一列在我的表格中是第“G”列
def getIP(wb):
    sheet = wb['Sheet1']
    ip = []
    for cellobj in sheet['A']:
        ip.append(cellobj.value)
    return ip

#获得某列数据
ip_a = getIP(wb_a)
ip_b = getIP(wb_b)

# 将两个列表转换成集合
aa = list(ip_a)
bb = list(ip_b)
print('777')
print(aa)
print(bb)
# tmp1=list(aa.union(bb))                 #并集
# tmp2=list(aa.intersection(bb))            # 交
# tmp4=list(aa.symmetric_difference(bb))    # 对称差
# 差  先转set集合，list集合没有difference属性
tmp6=list(set(aa).intersection(set(bb)))
print(tmp6)
#打印出列表中的元素
#到这一步，两个表格中相同的数据已经被找出来了
for i in tmp6:
    print(i)

 # list转dataframe
df = pd.DataFrame(tmp6, columns=['集合'])
df.to_excel('C:/Users/Administrator/Desktop/a.xlsx',index=False)


